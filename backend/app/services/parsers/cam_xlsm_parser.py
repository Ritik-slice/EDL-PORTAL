"""
Parser for .xlsm CAM Excel files from Slice's lending platform.

Extracts data from ALL 25 user-facing sheets plus the Output sheet.
Tracks editable (yellow-highlighted) cells for underwriter workflow.
Each sheet includes a ``_raw_grid`` for pixel-perfect frontend rendering.
"""

import io
import logging
from typing import Any, Optional

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Yellow / editable detection
# ---------------------------------------------------------------------------

_YELLOW_COLORS = {
    "FFFFFF00", "FFFBFF8E", "FFFFFF99",
    "FFFFFFCC", "FFFFF2CC", "FFFFFFE0", "FFFBFF00",
}


def _is_yellow(cell: Cell) -> bool:
    """Return True when *cell* has a yellow background (editable marker)."""
    try:
        fill = cell.fill
        for color_attr in ("fgColor", "bgColor"):
            color = getattr(fill, color_attr, None)
            if color and color.rgb:
                rgb = str(color.rgb).upper()
                if rgb in _YELLOW_COLORS:
                    return True
                # Heuristic: FFRRGGBB with high R, high G, low B
                if len(rgb) == 8 and rgb.startswith("FF"):
                    r = int(rgb[2:4], 16)
                    g = int(rgb[4:6], 16)
                    b = int(rgb[6:8], 16)
                    if r > 200 and g > 200 and b < 150:
                        return True
    except Exception:
        pass
    return False


# ---------------------------------------------------------------------------
# Value helpers
# ---------------------------------------------------------------------------

_FORMULA_ERRORS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")


def _safe_value(cell: Cell) -> Any:
    """Extract cell value, converting formula errors to None."""
    if cell is None:
        return None
    v = cell.value
    if isinstance(v, str) and any(v.startswith(e) for e in _FORMULA_ERRORS):
        return None
    return v


def _safe_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any) -> Optional[int]:
    f = _safe_float(val)
    return None if f is None else int(f)


# ---------------------------------------------------------------------------
# Sheet lookup
# ---------------------------------------------------------------------------

def _get_sheet(wb, name: str):
    """Case-insensitive sheet lookup. Returns None when missing."""
    lower = name.lower()
    for sn in wb.sheetnames:
        if sn.lower() == lower:
            return wb[sn]
    # Also try with common variations
    for sn in wb.sheetnames:
        sn_clean = sn.lower().replace(" ", "_").replace("-", "_")
        name_clean = lower.replace(" ", "_").replace("-", "_")
        if sn_clean == name_clean:
            return wb[sn]
    return None


def _find_sheet(wb_data, wb_style, name: str):
    """Return (ws_data, ws_style) tuple, either may be None."""
    return _get_sheet(wb_data, name), _get_sheet(wb_style, name)


# ---------------------------------------------------------------------------
# Raw grid builder
# ---------------------------------------------------------------------------

def _build_raw_grid(
    ws_data,
    ws_style,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    header_row: Optional[int] = None,
) -> dict:
    """
    Build a raw grid dict for the frontend.

    Returns::

        {
            "headers": [...],
            "rows": [
                {"cells": [{"value": ..., "editable": bool, "col": "A", "row": 1}, ...]}
            ]
        }
    """
    headers = []
    if header_row is not None and ws_data is not None:
        for col in range(min_col, max_col + 1):
            h = _safe_value(ws_data.cell(row=header_row, column=col))
            headers.append(_safe_str(h) if h is not None else None)

    rows = []
    if ws_data is None:
        return {"headers": headers, "rows": rows}

    for row in range(min_row, max_row + 1):
        cells = []
        for col in range(min_col, max_col + 1):
            val = _safe_value(ws_data.cell(row=row, column=col))
            editable = False
            if ws_style is not None:
                try:
                    editable = _is_yellow(ws_style.cell(row=row, column=col))
                except Exception:
                    pass
            cells.append({
                "value": val,
                "editable": editable,
                "col": get_column_letter(col),
                "row": row,
            })
        rows.append({"cells": cells})

    return {"headers": headers, "rows": rows}


# ---------------------------------------------------------------------------
# Editable-field tracker
# ---------------------------------------------------------------------------

def _track_if_yellow(
    editable_fields: list,
    ws_style,
    sheet_name: str,
    row: int,
    col: int,
    value: Any = None,
    key: Optional[str] = None,
):
    """Append to *editable_fields* when the style-workbook cell is yellow."""
    if ws_style is None:
        return
    try:
        cell = ws_style.cell(row=row, column=col)
        if _is_yellow(cell):
            editable_fields.append({
                "sheet": sheet_name,
                "cell": f"{get_column_letter(col)}{row}",
                "key": key,
                "current_value": value,
            })
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Generic helpers used by multiple sheet parsers
# ---------------------------------------------------------------------------

def _parse_label_value_sheet(
    ws_data, ws_style, sheet_name: str,
    label_col: int, value_col: int,
    min_row: int, max_row: int,
    editable_fields: list,
) -> dict:
    """Parse a sheet where col *label_col* has labels and *value_col* has values."""
    if ws_data is None:
        return {}
    result = {}
    for row in range(min_row, max_row + 1):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=label_col)))
        val = _safe_value(ws_data.cell(row=row, column=value_col))
        if label:
            result[label] = val
            _track_if_yellow(editable_fields, ws_style, sheet_name, row, value_col, val, label)
    return result


def _parse_table_sheet(
    ws_data, ws_style, sheet_name: str,
    header_row: int, data_min_row: int, data_max_row: int,
    min_col: int, max_col: int,
    editable_fields: list,
) -> list:
    """Parse a table sheet with a header row into a list of dicts."""
    if ws_data is None:
        return []

    headers = []
    for col in range(min_col, max_col + 1):
        h = _safe_str(_safe_value(ws_data.cell(row=header_row, column=col)))
        headers.append(h or f"col_{get_column_letter(col)}")

    items = []
    for row in range(data_min_row, data_max_row + 1):
        row_data = {}
        has_data = False
        for i, col in enumerate(range(min_col, max_col + 1)):
            val = _safe_value(ws_data.cell(row=row, column=col))
            if val is not None:
                has_data = True
            row_data[headers[i]] = val
            _track_if_yellow(editable_fields, ws_style, sheet_name, row, col, val, headers[i])
        if has_data:
            items.append(row_data)
    return items


# ---------------------------------------------------------------------------
# Sheet extent helpers
# ---------------------------------------------------------------------------

def _sheet_extent(ws) -> tuple:
    """Return (max_row, max_col) of actual data in the sheet."""
    if ws is None:
        return (0, 0)
    return (ws.max_row or 0, ws.max_column or 0)


# ===========================================================================
# Individual sheet parsers
# ===========================================================================

# 1. Demographic (108 rows) ------------------------------------------------

def _parse_demographic(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {}
    sn = "Demographic"
    max_row = max(108, _sheet_extent(ws_data)[0])
    max_col = min(10, _sheet_extent(ws_data)[1] or 10)

    result = {}
    current_section = "general"
    for row in range(1, max_row + 1):
        b = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        c = _safe_value(ws_data.cell(row=row, column=3))
        d = _safe_str(_safe_value(ws_data.cell(row=row, column=4)))
        e = _safe_value(ws_data.cell(row=row, column=5))

        # Section headers
        if b and c is None and d is None:
            lower = b.lower()
            if any(kw in lower for kw in [
                "sourcing", "loan app", "business", "co-borrower",
                "co-applicant", "co borrower", "co applicant",
            ]):
                current_section = b.strip()
                result.setdefault(current_section, {})
                continue

        section_dict = result.setdefault(current_section, {})
        if b and c is not None:
            section_dict[b] = c
            _track_if_yellow(editable_fields, ws_style, sn, row, 3, c, b)
        if d and e is not None:
            section_dict[d] = e
            _track_if_yellow(editable_fields, ws_style, sn, row, 5, e, d)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 2. Bureau (89 rows) ------------------------------------------------------

def _parse_bureau(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"scores": [], "existing_loans": [], "commercial_bureau": {}, "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Bureau"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(89, mr)
    max_col = max(17, mc)

    # Scores rows 3-15
    scores = []
    for row in range(3, 16):
        applicant_type = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        if not applicant_type:
            continue
        score_val = _safe_value(ws_data.cell(row=row, column=5))
        numeric_score = _safe_int(score_val)
        entry = {
            "applicant_index": _safe_value(ws_data.cell(row=row, column=1)),
            "type": applicant_type,
            "name": _safe_value(ws_data.cell(row=row, column=3)),
            "provider": _safe_str(_safe_value(ws_data.cell(row=row, column=4))),
            "score": numeric_score,
            "score_raw": score_val,
            "fetch_date": _safe_value(ws_data.cell(row=row, column=6)),
            "ntc_flag": _safe_value(ws_data.cell(row=row, column=7)),
        }
        scores.append(entry)
        for col in range(1, 8):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)))

    # Commercial bureau
    commercial_bureau = {
        "name": _safe_value(ws_data.cell(row=4, column=10)),
        "score": _safe_value(ws_data.cell(row=4, column=11)),
        "details": _safe_value(ws_data.cell(row=4, column=12)),
    }
    for col in [10, 11, 12]:
        _track_if_yellow(editable_fields, ws_style, sn, 4, col,
                         _safe_value(ws_data.cell(row=4, column=col)),
                         f"commercial_bureau_col{col}")

    # Existing loans rows 18+
    existing_loans = []
    for row in range(18, max_row + 1):
        applicant_idx = _safe_value(ws_data.cell(row=row, column=2))
        if applicant_idx is None:
            continue
        loan = {
            "applicant_index": applicant_idx,
            "borrower": _safe_value(ws_data.cell(row=row, column=3)),
            "financier": _safe_value(ws_data.cell(row=row, column=4)),
            "loan_type": _safe_value(ws_data.cell(row=row, column=5)),
            "amount": _safe_value(ws_data.cell(row=row, column=6)),
            "pos": _safe_value(ws_data.cell(row=row, column=7)),
            "term": _safe_value(ws_data.cell(row=row, column=8)),
            "mob": _safe_value(ws_data.cell(row=row, column=9)),
            "emi_assessed": _safe_value(ws_data.cell(row=row, column=10)),
            "obligated": _safe_value(ws_data.cell(row=row, column=12)),
            "bt_flag": _safe_value(ws_data.cell(row=row, column=14)),
            "foreclosure_flag": _safe_value(ws_data.cell(row=row, column=16)),
            "duplicate_flag": _safe_value(ws_data.cell(row=row, column=17)),
        }
        existing_loans.append(loan)
        for col in range(2, 18):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)),
                             f"existing_loan_r{row}_c{col}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=2)
    return {
        "scores": scores,
        "existing_loans": existing_loans,
        "commercial_bureau": commercial_bureau,
        "_raw_grid": raw,
    }


# 3. Banking (175 rows) ----------------------------------------------------

def _parse_banking(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"auto_pull": [], "manual_pull": [], "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Banking"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(175, mr)

    # Auto pull: cols B-P (2-16)
    auto_pull = []
    for row in range(1, max_row + 1):
        row_data = {}
        has_data = False
        for col in range(2, 17):  # B-P
            val = _safe_value(ws_data.cell(row=row, column=col))
            if val is not None:
                row_data[f"col_{get_column_letter(col)}"] = val
                has_data = True
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val)
        if has_data:
            auto_pull.append({"row": row, **row_data})

    # Manual pull: cols R-AG (18-33)
    manual_pull = []
    for row in range(1, max_row + 1):
        row_data = {}
        has_data = False
        for col in range(18, 34):  # R-AG
            val = _safe_value(ws_data.cell(row=row, column=col))
            if val is not None:
                row_data[f"col_{get_column_letter(col)}"] = val
                has_data = True
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val)
        if has_data:
            manual_pull.append({"row": row, **row_data})

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, min(33, mc or 33), header_row=1)
    return {"auto_pull": auto_pull, "manual_pull": manual_pull, "_raw_grid": raw}


# 4. Additional Credit Checks (16 rows) ------------------------------------

def _parse_additional_credit_checks(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Additional Credit Checks"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(16, mr)
    max_col = max(8, mc or 8)

    result = {}
    for row in range(1, max_row + 1):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        val = _safe_value(ws_data.cell(row=row, column=3))
        if label:
            result[label] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, 3, val, label)
        # Some checks have additional columns
        for col in range(4, max_col + 1):
            extra_val = _safe_value(ws_data.cell(row=row, column=col))
            if extra_val is not None and label:
                result[f"{label}_col{get_column_letter(col)}"] = extra_val
                _track_if_yellow(editable_fields, ws_style, sn, row, col, extra_val, f"{label}_col{get_column_letter(col)}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 5. GST (30 rows) ---------------------------------------------------------

def _parse_gst(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"auto_pull": {}, "manual_pull": {}, "monthly_filings": [],
                "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "GST"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(30, mr)

    # Auto pull header: rows 3-6, B=label, C=value, D=extra
    auto_pull = {}
    for row in [3, 4, 5, 6]:
        key = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        val = _safe_value(ws_data.cell(row=row, column=3))
        if key:
            auto_pull[key] = val

    # Auto pull monthly: rows 10-21, B=Month, C=FY, D=Amount
    auto_monthly = []
    for row in range(10, 22):
        month = _safe_value(ws_data.cell(row=row, column=2))
        if month is None or str(month).strip() in ("", "Total", "00:00:00"):
            continue
        auto_monthly.append({
            "month": month,
            "fy": _safe_value(ws_data.cell(row=row, column=3)),
            "amount": _safe_value(ws_data.cell(row=row, column=4)),
            "source": "auto",
        })

    # Manual pull header: rows 3-6, H=label, I=value
    manual_pull = {}
    manual_field_map = {3: "gstin", 4: "status", 5: "latest_month", 6: "frequency"}
    for row, field_name in manual_field_map.items():
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=8)))
        val = _safe_value(ws_data.cell(row=row, column=9))
        manual_pull[field_name] = val
        _track_if_yellow(editable_fields, ws_style, sn, row, 9, val, f"gst_{field_name}")

    # Manual pull monthly: rows 10-21, H=Month, I=FY, J=Amount
    manual_monthly = []
    for row in range(10, 22):
        month = _safe_value(ws_data.cell(row=row, column=8))
        if month is None or str(month).strip() in ("", "Total", "00:00:00"):
            continue
        manual_monthly.append({
            "month": month,
            "fy": _safe_value(ws_data.cell(row=row, column=9)),
            "amount": _safe_value(ws_data.cell(row=row, column=10)),
            "source": "manual",
            "editable": True,
        })
        for c in [8, 9, 10]:
            _track_if_yellow(editable_fields, ws_style, sn, row, c,
                             _safe_value(ws_data.cell(row=row, column=c)), f"gst_manual_r{row}")

    monthly_filings = auto_monthly + manual_monthly

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, min(10, mc or 10), header_row=None)
    return {
        "auto_pull": auto_pull,
        "manual_pull": manual_pull,
        "monthly_filings": monthly_filings,
        "_raw_grid": raw,
    }


# 6. ITR (5 rows) ----------------------------------------------------------

def _parse_itr(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "ITR"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(5, mr)
    max_col = max(6, mc or 6)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 7. Company Financials (39 rows) ------------------------------------------

def _parse_company_financials(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"balance_sheet": {}, "pnl": {}, "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Company Financials"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(39, mr)
    max_col = max(6, mc or 6)

    # Detect year headers from row 7 or row 1
    year_headers = {}
    for col in range(3, max_col + 1):
        for hdr_row in [1, 7]:
            h = _safe_str(_safe_value(ws_data.cell(row=hdr_row, column=col)))
            if h:
                year_headers[col] = h
                break
        if col not in year_headers:
            year_headers[col] = f"col_{get_column_letter(col)}"

    # Balance Sheet rows 8-16
    balance_sheet = {}
    for row in range(8, 17):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        if not label:
            continue
        entry = {}
        for col in range(3, max_col + 1):
            val = _safe_value(ws_data.cell(row=row, column=col))
            entry[year_headers.get(col, f"col_{col}")] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val, label)
        balance_sheet[label] = entry

    # P&L rows 17-35
    pnl = {}
    for row in range(17, 36):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        if not label:
            continue
        entry = {}
        for col in range(3, max_col + 1):
            val = _safe_value(ws_data.cell(row=row, column=col))
            entry[year_headers.get(col, f"col_{col}")] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val, label)
        pnl[label] = entry

    # Also grab remaining rows (36-39) as extras
    extras = {}
    for row in range(36, max_row + 1):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        if not label:
            continue
        entry = {}
        for col in range(3, max_col + 1):
            val = _safe_value(ws_data.cell(row=row, column=col))
            entry[year_headers.get(col, f"col_{col}")] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val, label)
        extras[label] = entry

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=7)
    return {"balance_sheet": balance_sheet, "pnl": pnl, "extras": extras, "_raw_grid": raw}


# 8. AIP (27 rows) ---------------------------------------------------------

def _parse_aip(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"primary_income": {}, "bills_collection": [], "secondary_income": [],
                "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "AIP"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(27, mr)

    # Primary income: cols B-G (2-7)
    primary_income = {}
    for row in range(2, max_row + 1):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        if not label:
            continue
        entry = {}
        for col in range(3, 8):  # C-G
            header = _safe_str(_safe_value(ws_data.cell(row=1, column=col))) or f"col_{get_column_letter(col)}"
            val = _safe_value(ws_data.cell(row=row, column=col))
            entry[header] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val, label)
        primary_income[label] = entry

    # Bills collection: cols J-O (10-15)
    bills_collection = []
    for row in range(2, max_row + 1):
        bill_type = _safe_value(ws_data.cell(row=row, column=10))
        if bill_type is None:
            continue
        bill = {"type": bill_type}
        for col in range(11, 16):  # K-O
            header = _safe_str(_safe_value(ws_data.cell(row=1, column=col))) or f"col_{get_column_letter(col)}"
            val = _safe_value(ws_data.cell(row=row, column=col))
            bill[header] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, col, val, f"bill_r{row}")
        bills_collection.append(bill)

    # Secondary income (after primary, further rows)
    secondary_income = []
    for row in range(2, max_row + 1):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        val = _safe_value(ws_data.cell(row=row, column=3))
        if label and val is not None and "secondary" in label.lower():
            secondary_income.append({"label": label, "value": val})

    max_c = max(15, mc or 15)
    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_c, header_row=1)
    return {
        "primary_income": primary_income,
        "bills_collection": bills_collection,
        "secondary_income": secondary_income,
        "_raw_grid": raw,
    }


# 9. PD notes (20 rows) ----------------------------------------------------

def _parse_pd_notes(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "PD notes"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(20, mr)
    max_col = max(5, mc or 5)

    result = {}
    current_section = None

    for row in range(1, max_row + 1):
        col_a = _safe_str(_safe_value(ws_data.cell(row=row, column=1)))
        col_d = _safe_value(ws_data.cell(row=row, column=4))

        # Section headers
        if col_a and "notes on" in col_a.lower():
            current_section = col_a
            continue

        # Free-text lines under a "Notes on" section
        if current_section and col_a and col_d is None:
            if current_section not in result:
                result[current_section] = col_a
            else:
                result[current_section] += "\n" + col_a
            _track_if_yellow(editable_fields, ws_style, sn, row, 1,
                             col_a, current_section)
            continue

        # Structured: col A = label, col D = value
        if col_a and col_d is not None:
            result[col_a] = col_d
            _track_if_yellow(editable_fields, ws_style, sn, row, 4, col_d, col_a)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col)
    result["_raw_grid"] = raw
    return result


# 10. Reference checks (4 rows) --------------------------------------------

def _parse_reference_checks(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"references": [], "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Reference checks"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(4, mr)
    max_col = max(6, mc or 6)

    references = []
    for row in range(2, max_row + 1):
        name = _safe_value(ws_data.cell(row=row, column=2))
        if name is None:
            continue
        ref = {
            "name": name,
            "relation": _safe_value(ws_data.cell(row=row, column=3)),
            "phone": _safe_value(ws_data.cell(row=row, column=4)),
            "feedback": _safe_value(ws_data.cell(row=row, column=5)),
        }
        references.append(ref)
        for col in range(2, max_col + 1):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)),
                             f"ref_check_r{row}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    return {"references": references, "_raw_grid": raw}


# 11. Loan purpose (12 rows) -----------------------------------------------

def _parse_loan_purpose(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"items": [], "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Loan purpose"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(12, mr)
    max_col = max(5, mc or 5)

    items = []
    for row in range(2, max_row + 1):
        sno = _safe_value(ws_data.cell(row=row, column=1))
        end_use = _safe_value(ws_data.cell(row=row, column=2))
        if sno is None and end_use is None:
            continue
        item = {
            "s_no": sno,
            "end_use": end_use,
            "details": _safe_value(ws_data.cell(row=row, column=3)),
            "amount": _safe_value(ws_data.cell(row=row, column=4)),
        }
        items.append(item)
        for col in range(1, max_col + 1):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)),
                             f"loan_purpose_r{row}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    return {"items": items, "_raw_grid": raw}


# 12. Stock Details (11 rows) ----------------------------------------------

def _parse_stock_details(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"items": [], "total": None, "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Stock Details"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(23, mr)
    max_col = max(5, mc or 5)

    items = []
    total = None
    for row in range(2, max_row + 1):
        sl_no = _safe_value(ws_data.cell(row=row, column=1))
        desc = _safe_value(ws_data.cell(row=row, column=2))

        # Row 23 is total
        if row == 23 or (sl_no is not None and str(sl_no).strip().lower() == "total"):
            total = _safe_value(ws_data.cell(row=row, column=5))
            continue

        if sl_no is None and desc is None:
            continue

        item = {
            "sl_no": sl_no,
            "description": desc,
            "quantity": _safe_value(ws_data.cell(row=row, column=3)),
            "rate": _safe_value(ws_data.cell(row=row, column=4)),
            "amount": _safe_value(ws_data.cell(row=row, column=5)),
        }
        items.append(item)
        for col in range(1, 6):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)),
                             f"stock_r{row}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    return {"items": items, "total": total, "_raw_grid": raw}


# 13. Plant & Machinery Details (23 rows) ----------------------------------

def _parse_plant_machinery(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"movable_pm": [], "vehicles": [],
                "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Plant & Machinery Details"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(23, mr)
    max_col = max(8, mc or 8)

    # Headers from row 1
    headers = {}
    for col in range(1, max_col + 1):
        h = _safe_str(_safe_value(ws_data.cell(row=1, column=col)))
        headers[col] = h or f"col_{get_column_letter(col)}"

    def _parse_rows(start: int, end: int) -> list:
        items = []
        for row in range(start, end + 1):
            row_data = {}
            has_data = False
            for col in range(1, max_col + 1):
                val = _safe_value(ws_data.cell(row=row, column=col))
                if val is not None:
                    has_data = True
                row_data[headers[col]] = val
                _track_if_yellow(editable_fields, ws_style, sn, row, col, val,
                                 f"pm_r{row}")
            if has_data:
                items.append(row_data)
        return items

    movable_pm = _parse_rows(2, 12)
    vehicles = _parse_rows(14, 22)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    return {"movable_pm": movable_pm, "vehicles": vehicles, "_raw_grid": raw}


# 14. Property Details (98 rows) -------------------------------------------

def _parse_property_details(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Property Details"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(98, mr)
    max_col = max(8, mc or 8)

    result = {}
    current_section = "general"
    for row in range(1, max_row + 1):
        b = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        c = _safe_value(ws_data.cell(row=row, column=3))

        # Detect section headers (Valuer 1, Valuer 2, LTV, Mortgager, etc.)
        if b and c is None:
            lower = b.lower()
            if any(kw in lower for kw in [
                "valuer", "ltv", "property", "mortgager", "collateral",
                "market value", "distress",
            ]):
                current_section = b.strip()
                result.setdefault(current_section, {})
                continue

        section_dict = result.setdefault(current_section, {})
        if b and c is not None:
            section_dict[b] = c
            _track_if_yellow(editable_fields, ws_style, sn, row, 3, c, b)

        # Also capture cols D-E if present
        d = _safe_str(_safe_value(ws_data.cell(row=row, column=4)))
        e = _safe_value(ws_data.cell(row=row, column=5))
        if d and e is not None:
            section_dict[d] = e
            _track_if_yellow(editable_fields, ws_style, sn, row, 5, e, d)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 15. Scorecard (38 rows) --------------------------------------------------

def _parse_scorecard(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"total_score": None, "risk_band": None, "parameters": [],
                "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Scorecard"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(38, mr)
    max_col = max(8, mc or 8)

    # Total Score at B3/D3, Risk Band at B4/D4
    total_score = _safe_value(ws_data.cell(row=3, column=4))
    if total_score is None:
        total_score = _safe_value(ws_data.cell(row=3, column=2))
    total_score = _safe_int(total_score)

    risk_band = _safe_str(_safe_value(ws_data.cell(row=4, column=4)))
    if not risk_band:
        risk_band = _safe_str(_safe_value(ws_data.cell(row=4, column=2)))

    # Parameters rows 8+: B=sr_no, C=parameter, F=max_score, G=criteria(YELLOW), H=marks
    parameters = []
    for row in range(8, max_row + 1):
        param_name = _safe_str(_safe_value(ws_data.cell(row=row, column=3)))
        marks = _safe_value(ws_data.cell(row=row, column=8))
        if not param_name:
            continue
        param = {
            "sr_no": _safe_value(ws_data.cell(row=row, column=2)),
            "parameter": param_name,
            "max_score": _safe_value(ws_data.cell(row=row, column=6)),
            "criteria": _safe_value(ws_data.cell(row=row, column=7)),
            "marks": marks,
        }
        parameters.append(param)
        _track_if_yellow(editable_fields, ws_style, sn, row, 7,
                         _safe_value(ws_data.cell(row=row, column=7)),
                         f"scorecard_criteria_{param_name}")
        _track_if_yellow(editable_fields, ws_style, sn, row, 8, marks,
                         f"scorecard_marks_{param_name}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=7)
    return {
        "total_score": total_score,
        "risk_band": risk_band,
        "parameters": parameters,
        "_raw_grid": raw,
    }


# 16. Deviations_Credit (40 rows) ------------------------------------------

def _parse_deviations_credit(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"deviations": [], "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Deviations_Credit"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(40, mr)
    max_col = max(8, mc or 8)

    deviations = []
    for row in range(2, max_row + 1):
        dev = _safe_value(ws_data.cell(row=row, column=2))
        if dev is None:
            continue
        entry = {
            "deviation": dev,
            "description": _safe_value(ws_data.cell(row=row, column=3)),
            "mitigants": _safe_value(ws_data.cell(row=row, column=4)),
            "authority": _safe_value(ws_data.cell(row=row, column=5)),
            "status": _safe_value(ws_data.cell(row=row, column=6)),
            "approved_by": _safe_value(ws_data.cell(row=row, column=7)),
            "date": _safe_value(ws_data.cell(row=row, column=8)),
        }
        deviations.append(entry)
        for col in range(2, max_col + 1):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)),
                             f"deviation_r{row}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    return {"deviations": deviations, "_raw_grid": raw}


# 17. Interest Rate Calculator (8 rows) ------------------------------------

def _parse_interest_rate(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Interest Rate Calculator"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(8, mr)
    max_col = max(4, mc or 4)

    fields = [
        (2, "security_type"),
        (3, "loan_band"),
        (4, "geography"),
        (5, "bureau_band"),
        (6, "program"),
        (7, "deviations"),
        (8, "final_rate"),
    ]
    result = {}
    for row, key in fields:
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        val = _safe_value(ws_data.cell(row=row, column=3))
        result[key] = val
        _track_if_yellow(editable_fields, ws_style, sn, row, 3, val, key)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 18. Eligibility - Inventory (67 rows) ------------------------------------

def _parse_eligibility_inventory(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Eligibility - Inventory"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(67, mr)
    max_col = max(6, mc or 6)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 19. Eligibility - Property (67 rows) -------------------------------------

def _parse_eligibility_property(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Eligibility - Property"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(67, mr)
    max_col = max(6, mc or 6)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 20. Sanction Conditions (16 rows) ----------------------------------------

def _parse_sanction_conditions(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"conditions": [], "_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Sanction Conditions"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(16, mr)
    max_col = max(4, mc or 4)

    conditions = []
    for row in range(2, max_row + 1):
        val = _safe_value(ws_data.cell(row=row, column=2))
        if val is None:
            val = _safe_value(ws_data.cell(row=row, column=1))
        if val is None:
            continue
        conditions.append(val)
        for col in range(1, max_col + 1):
            _track_if_yellow(editable_fields, ws_style, sn, row, col,
                             _safe_value(ws_data.cell(row=row, column=col)),
                             f"sanction_r{row}")

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    return {"conditions": conditions, "_raw_grid": raw}


# 21. PSL (8 rows) ---------------------------------------------------------

def _parse_psl(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "PSL"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(8, mr)
    max_col = max(4, mc or 4)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    # Also try col A=label, col B=value
    if not result:
        result = _parse_label_value_sheet(ws_data, ws_style, sn, 1, 2, 1, max_row, editable_fields)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 22. CGTMSE Calculator (17 rows) ------------------------------------------

def _parse_cgtmse(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "CGTMSE Calculator"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(17, mr)
    max_col = max(4, mc or 4)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    # Also try col A=label, col B=value
    extra = _parse_label_value_sheet(ws_data, ws_style, sn, 1, 2, 1, max_row, editable_fields)
    for k, v in extra.items():
        if k not in result:
            result[k] = v

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 23. Charges (21 rows) ----------------------------------------------------

def _parse_charges(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Charges"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(21, mr)
    max_col = max(5, mc or 5)

    result = {}
    for row in range(1, max_row + 1):
        label = _safe_str(_safe_value(ws_data.cell(row=row, column=2)))
        if not label:
            label = _safe_str(_safe_value(ws_data.cell(row=row, column=1)))
        val = _safe_value(ws_data.cell(row=row, column=3))
        if val is None:
            val = _safe_value(ws_data.cell(row=row, column=2)) if label == _safe_str(_safe_value(ws_data.cell(row=row, column=1))) else None
        if label and val is not None:
            result[label] = val
            _track_if_yellow(editable_fields, ws_style, sn, row, 3, val, label)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 24. Disbursement (5 rows) ------------------------------------------------

def _parse_disbursement(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Disbursement"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(5, mr)
    max_col = max(5, mc or 5)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    if not result:
        result = _parse_label_value_sheet(ws_data, ws_style, sn, 1, 2, 1, max_row, editable_fields)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# 25. Additional Information (12 rows) -------------------------------------

def _parse_additional_information(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {"_raw_grid": _build_raw_grid(None, None, 1, 1, 1, 1)}
    sn = "Additional Information"
    mr, mc = _sheet_extent(ws_data)
    max_row = max(12, mr)
    max_col = max(5, mc or 5)

    result = _parse_label_value_sheet(ws_data, ws_style, sn, 2, 3, 1, max_row, editable_fields)
    if not result:
        result = _parse_label_value_sheet(ws_data, ws_style, sn, 1, 2, 1, max_row, editable_fields)

    raw = _build_raw_grid(ws_data, ws_style, 1, max_row, 1, max_col, header_row=1)
    result["_raw_grid"] = raw
    return result


# ---------------------------------------------------------------------------
# Output sheet (160 key-value pairs)
# ---------------------------------------------------------------------------

def _parse_output(ws_data, ws_style, editable_fields: list) -> dict:
    if ws_data is None:
        return {}
    sn = "Output"
    mr = max(170, _sheet_extent(ws_data)[0])

    result = {}
    for row in range(1, mr + 1):
        key = _safe_str(_safe_value(ws_data.cell(row=row, column=1)))
        value = _safe_value(ws_data.cell(row=row, column=2))
        if key:
            result[key] = value
            _track_if_yellow(editable_fields, ws_style, sn, row, 2, value, key)
    return result


# ===========================================================================
# Sheet name mapping: user-facing name -> (snake_case key, parser function)
# ===========================================================================

# Names to try for each sheet (first match wins)
_SHEET_CONFIGS = [
    # (result_key, [sheet_name_variants], parser_func)
    ("demographic", ["Demographic"], _parse_demographic),
    ("bureau", ["Bureau"], _parse_bureau),
    ("banking", ["Banking"], _parse_banking),
    ("additional_credit_checks", ["Additional Credit Checks"], _parse_additional_credit_checks),
    ("gst", ["GST"], _parse_gst),
    ("itr", ["ITR"], _parse_itr),
    ("company_financials", ["Company Financials"], _parse_company_financials),
    ("aip", ["AIP"], _parse_aip),
    ("pd_notes", ["PD notes", "PD Notes"], _parse_pd_notes),
    ("reference_checks", ["Reference checks", "Reference Checks"], _parse_reference_checks),
    ("loan_purpose", ["Loan purpose", "Loan Purpose"], _parse_loan_purpose),
    ("stock_details", ["Stock Details"], _parse_stock_details),
    ("plant_machinery_details", ["Plant & Machinery Details", "Plant and Machinery Details", "Plant & Machinery"], _parse_plant_machinery),
    ("property_details", ["Property Details"], _parse_property_details),
    ("scorecard", ["Scorecard"], _parse_scorecard),
    ("deviations_credit", ["Deviations_Credit", "Deviations Credit"], _parse_deviations_credit),
    ("interest_rate_calculator", ["Interest Rate Calculator"], _parse_interest_rate),
    ("eligibility_inventory", ["Eligibility - Inventory", "Eligibility_Inventory", "Eligibility Inventory"], _parse_eligibility_inventory),
    ("eligibility_property", ["Eligibility - Property", "Eligibility_Property", "Eligibility Property"], _parse_eligibility_property),
    ("sanction_conditions", ["Sanction Conditions"], _parse_sanction_conditions),
    ("psl", ["PSL"], _parse_psl),
    ("cgtmse_calculator", ["CGTMSE Calculator", "CGTMSE"], _parse_cgtmse),
    ("charges", ["Charges"], _parse_charges),
    ("disbursement", ["Disbursement"], _parse_disbursement),
    ("additional_information", ["Additional Information"], _parse_additional_information),
]


# ===========================================================================
# Main entry point
# ===========================================================================

def parse_cam_xlsm(content: bytes, filename: str) -> dict:
    """
    Parse a .xlsm CAM Excel file and extract structured data from all 25
    user-facing sheets plus the Output sheet.

    Args:
        content: Raw bytes of the .xlsm file.
        filename: Original filename (for logging).

    Returns:
        dict with a top-level key per sheet (snake_case), plus ``editable_fields``
        list and ``output`` dict.
    """
    logger.info("Parsing CAM XLSM file: %s (%d bytes)", filename, len(content))

    # Load workbook twice:
    #   wb_data  – data_only=True  → computed values (no formulas)
    #   wb_style – data_only=False → cell styles / colors
    try:
        wb_data = openpyxl.load_workbook(
            io.BytesIO(content), data_only=True, keep_vba=True, read_only=False,
        )
    except Exception as e:
        logger.error("Failed to load workbook (data_only): %s", e)
        raise ValueError(f"Failed to parse XLSM file '{filename}': {e}") from e

    try:
        wb_style = openpyxl.load_workbook(
            io.BytesIO(content), data_only=False, keep_vba=True, read_only=False,
        )
    except Exception:
        logger.warning("Could not load style workbook; editable field detection may be limited")
        wb_style = wb_data

    logger.info("Sheets found: %s", wb_data.sheetnames)

    editable_fields: list[dict] = []
    result: dict[str, Any] = {}

    # Parse each of the 25 user-facing sheets
    for key, name_variants, parser_fn in _SHEET_CONFIGS:
        ws_data = None
        ws_style_sheet = None
        for name in name_variants:
            ws_data = _get_sheet(wb_data, name)
            ws_style_sheet = _get_sheet(wb_style, name)
            if ws_data is not None:
                break

        if ws_data is None:
            logger.warning("Sheet not found for key '%s' (tried: %s)", key, name_variants)

        try:
            result[key] = parser_fn(ws_data, ws_style_sheet, editable_fields)
        except Exception:
            logger.exception("Error parsing sheet '%s'", key)
            result[key] = {}

    # Output sheet (special: 160 key-value pairs)
    ws_out_data = _get_sheet(wb_data, "Output")
    ws_out_style = _get_sheet(wb_style, "Output")
    result["output"] = _parse_output(ws_out_data, ws_out_style, editable_fields)

    result["editable_fields"] = editable_fields

    # Cleanup
    try:
        wb_data.close()
        if wb_style is not wb_data:
            wb_style.close()
    except Exception:
        pass

    logger.info(
        "Parsed %s: %d sheets extracted, %d output keys, %d editable fields",
        filename,
        sum(1 for k, v in result.items() if k not in ("editable_fields", "output") and v),
        len(result.get("output", {})),
        len(editable_fields),
    )

    return result
